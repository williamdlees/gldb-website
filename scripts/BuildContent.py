# Copyright (c) 2018 William Lees

# Permission is hereby granted, free of charge, to any person obtaining a copy of this software and associated
# documentation files (the "Software"), to deal in the Software without restriction, including without limitation the
# rights to use, copy, modify, merge, publish, distribute, sublicense, and/or sell copies of the Software, and to permit
# persons to whom the Software is furnished to do so, subject to the following conditions:

# The above copyright notice and this permission notice shall be included in all copies or substantial portions of the
# Software.

# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE
# WARRANTIES OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE AUTHORS OR
# COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR
# OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.

__author__ = 'William Lees'
__docformat__ = "restructuredtext en"

# Script to extract all info from IARC sheets and minutes, and to create from it the files required by hugo to build the static website


import os
import sys
import shutil
import itertools
import json
from collections import OrderedDict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, GradientFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import coordinate_from_string, column_index_from_string
from Bio.Seq import Seq
from Bio.Alphabet import IUPAC
import itertools as itertools



# Locations of the 'base' tree, containing content in excel, pdf format

base = 'base'
layout = 'build/layouts'
content = 'build/content'
static = 'build/static'

# Globals that are populated from the base content

minute_details = {}
submission_details = {}
sequence_details = {}

def main(argv):
    # Extract data from source files, and copy source files into static
    extract_data_from_minutes()
    extract_data_from_submissions()
    extract_data_from_sequences()

    # Create hugo front matter
    create_meeting_content()
    create_submissions_content()
    create_sequences_content()


# 'Business Layer' functions that build data required by the site from base content

def create_meeting_content():
    makempty(content + '/meetings')
    for (id, rec) in minute_details.items():
        with open('%s/meetings/%s' % (content, rec['filename'].replace('.pdf', '.md')), 'w', encoding='utf-8') as fo:
            fo.write('---\n')
            fo.write('number: %d\n' % id)
            fo.write('date: %s\n' % rec['date'])
            fo.write('url: %s\n' % rec['url'])
            fo.write('---\n')

def create_submissions_content():
    makempty(content + '/submissions')
    for(id, rec) in submission_details.items():
        with open('%s/submissions/%s' % (content, rec['filename'].replace('.xlsx', '.md')), 'w', encoding='utf-8') as fo:
            data = rec['data']

            data['file_url'] = rec['url']
            data['sup_url'] = rec['sup_url']

            # turn URLs into clickable links
            pmids = str((int(data['Repertoire']['pub_ids']))).split(',')

            data['Repertoire']['pub_ids'] = ''
            pmidl = []
            for pmid in pmids:
                pmid = pmid.replace(' ', '')
                url = 'https://www.ncbi.nlm.nih.gov/pubmed/?term=' + pmid
                pmidl.append(addressify(url, pmid))
            data['Repertoire']['pub_ids']  = ','.join(pmidl)

            data['Repertoire']['dataset_url'] = addressify(data['Repertoire']['dataset_url'], data['Repertoire']['dataset_url'])
            data['Repertoire']['dataset_doi'] = addressify(data['Repertoire']['dataset_doi'], data['Repertoire']['dataset_doi'])

            # Notes tab

            data['IARC_tracking'] = data['Submission']['IARC_tracking'].split("\n")
            del data['Submission']['IARC_tracking']

            # build lists that allow fields to be displayed in the correct order
            data['Submission_order'] = list(data['Submission'].keys())
            data['Repertoire_order'] = list(data['Repertoire'].keys())
            first = data['Inference_list'][0]
            data['Inference_head'] = list(data['Inferences'][first][0].keys())
            first = data['Genotype_list'][0]
            data['Genotype_head'] = list(data['Genotype'][first][0].keys())
            data['Tool_Settings_order'] = list(data['Tool_Settings'].keys())


            fo.write(json.dumps({'date': rec['date'], 'content': data}, indent=4, sort_keys=False, default=str))

def create_sequences_content():
    makempty(content + '/sequences')

    for(id, rec) in sequence_details.items():
        with open('%s/sequences/%s' % (content, rec['filename'].replace('.xlsx', '.md')), 'w', encoding='utf-8') as fo:
            data = rec['data']

            data['file_url'] = rec['url']

            subs = data['Submissions']
            data['Submissions'] = []
            for sub, gene in subs.items():
                data['Submissions'].append({'id': str(sub), 'url': 'submissions/iarc_submission_%s' % str(sub), 'name': gene})

            mtgs = data['Meetings']
            data['Meetings'] = []
            for mtg in mtgs:
                if mtg in minute_details:
                    data['Meetings'].append(minute_details[mtg])
                else:
                    print('Meeting %s, referred to in filename %s, does not exist in the system.' % (mtg, rec['filename']))

            # Pubmed ids
            pmids = str((int(data['Sequence']['pub_ids']))).split(',')

            data['Sequence']['pub_ids'] = ''
            pmidl = []
            for pmid in pmids:
                pmid = pmid.replace(' ', '')
                url = 'https://www.ncbi.nlm.nih.gov/pubmed/?term=' + pmid
                pmidl.append(addressify(url, pmid))
            data['Sequence']['pub_ids']  = ','.join(pmidl)

            # Hide non-coding regions if there are no genomic sequences

            if len(str(data['Sequence']['genomic_sequences']).replace(' ', '')) == 0:
                for field in ['utr_5_prime', 'l_region', 'v_rs', 'd_rs_3_prime', 'd_rs_5_prime', 'j_rs']:
                    del data['Sequence'][field + '_start']
                    del data['Sequence'][field + '_end']

            # Notes tab

            data['Notes'] = data['Sequence']['notes'].split("\n")
            del data['Sequence']['notes']

            #  Sequences formatted for display

            data['fmt_raw'] = format_nuc_sequence(data['Sequence']['sequence'], 50)
            data['fmt_fa'] = format_fasta_sequence(data['Sequence']['gene_name'], data['Sequence']['sequence'], 50)

            if 'V' in data['Sequence']['region']:
                data['fmt_imgt'] = format_imgt_v(data['Sequence']['coding_seq_imgt'], 100)
            elif len(data['Sequence']['sequence']) != len(data['Sequence']['coding_seq_imgt']):
                data['fmt_imgt'] = format_nuc_sequence(data['Sequence']['coding_seq_imgt'], 50)
            else:
                data['fmt_imgt'] = ''

            del data['Sequence']['sequence']
            del data['Sequence']['coding_seq_imgt']

            # build lists that allow fields to be displayed in the correct order
            data['Sequence_order'] = list(data['Sequence'].keys())
            fo.write(json.dumps({'date': rec['date'], 'file_url': rec['url'], 'content': rec['data']}, indent=4, sort_keys=False, default=str))

def addressify(url, name):
    return '<a href="%s" target="_blank">%s</a>' % (url, name)

def chunks(l, n):
    " Yield successive n-sized chunks from l."
    for i in range(0, len(l), n):
        yield l[i:i + n]

def format_nuc_sequence(seq, width):
    ind = 1
    ret = ''

    for frag in chunks(seq, width):
        ret += '%-5d' % ind
        if len(frag) > 10:
            ret += ' '*(len(frag)-10) + '%5d' % (ind + len(frag) - 1)
        ind += len(frag)
        ret += '\n' + frag + '\n\n'

    return ret

def format_fasta_sequence(name, seq, width):
    ret = '>' + name + '\n'

    for frag in chunks(seq, width):
        ret += frag + '\n'

    return ret

imgt_leg = '                                                                                                       _____________________CDR1_______________________                                                                     _________________CDR2___________________                                                                                                                                                             _CDR3_______'
imgt_num = ' 1   2   3   4   5   6   7   8   9   10  11  12  13  14  15  16  17  18  19  20  21  22  23  24  25  26  27  28  29  30  31  32  33  34  35  36  37  38  39  40  41  42  43  44  45  46  47  48  49  50  51  52  53  54  55  56  57  58  59  60  61  62  63  64  65  66  67  68  69  70  71  72  73  74  75  76  77  78  79  80  81  82  83  84  85  86  87  88  89  90  91  92  93  94  95  96  97  98  99 100 101 102 103 104 105 106     '

def format_imgt_v(seq, width):
    ind = 1
    ret = 1

    fmt_seq = ''
    fmt_aa = ''
    for cd in chunks(seq, 3):
        fmt_seq += cd + ' '
        if '.' in cd or len(cd) < 3:
            fmt_aa += '    '
        else:
            fmt_aa += ' ' + str(Seq(cd, IUPAC.unambiguous_dna).translate()) + '  '

    # this will deliberately truncate at the end of the shortest line - which will never be imgt_leg or imgt_num unless the sequence is longer than it should be...
    return splitlines(imgt_leg + '\n' + imgt_num + '\n' + fmt_aa + '\n' + fmt_seq + '\n', width, 0)

def splitlines(report, maxlength, label_cols):
    """
    Split the report (which is assumed to consist of lines of equal length) into a longer report in which each
    line is maxlength or less. name_cols specifies the width of the label field, which is repeated at the start
    of each line.
    """

    # https://stackoverflow.com/questions/3992735/python-generator-that-groups-another-iterable-into-groups-of-n

    def grouper(n, iterable):
        iterable = iter(iterable)
        return iter(lambda: list(itertools.islice(iterable, n)), [])

    inlines = report.split("\n")[:-1]
    labels = [line[:label_cols] for line in inlines]
    data = [line[label_cols:] for line in inlines]
    outlines = []

    for chunk in grouper(maxlength-label_cols, zip(*data)):
        a = ["".join(line) for line in zip(*chunk)]
        outlines.extend(["".join(line) for line in zip(labels, a)])
        outlines.extend(" ")

    return "\n".join(["".join(line) for line in outlines])


# 'Extraction Layer' functions that extract data from base content, and move base content files into static

def extract_data_from_minutes():
    global minute_details

    makempty(static + '/meetings')

    # Filename format must be 'IARC Meeting <meetingno> <day> <month> <year> minutes.pdf'
    # We allow underlines instead of spaces
    # We'll try to tolerate typos but the number of words must agree
    minutes = get_files_in_dir(base + '/meetings', '.pdf')
    for mn in minutes:
        fn = mn.replace(' ', '_')
        while('__') in fn:
            fn = fn.replace('__', '_')
        targname = fn.lower()
        fn = fn.split('_')
        if len(fn) != 7:
            print('%s/meetings/%s: filename does not follow correct format.' % (base, mn))
            continue
        try:
            meetingnum = int(fn[2])
            date = '%4d-%02d-%02d' % ((int(fn[5]) if int(fn[5]) > 100 else 2000 + int(fn[5])), int(fn[4]), int(fn[3]))
        except ValueError:
            print('%s/meetings/%s: filename does not follow correct format.' % (base, mn))
            continue

        shutil.copyfile(base + '/meetings/' + mn, static + '/meetings/' + targname)
        minute_details[meetingnum] = {'id': meetingnum, 'url': 'meetings/%s' % targname, 'date': date, 'filename': targname}

def extract_data_from_submissions():
    global submission_details

    makempty(static + '/submissions')

    # Filename format must be 'iarc_submission_<number>.xlsx'
    subs = get_files_in_dir(base + '/submissions', '.xlsx')
    for sub in subs:
        if 'supplementary' not in sub:
            fn = sub.replace(' ', '_')
            while('__') in fn:
                fn = fn.replace('__', '_')
            targname = fn.lower()
            fn = fn.replace('.xlsx', '')
            fn = fn.split('_')
            if len(fn) != 3:
                print('%s/submissions/%s: filename does not follow correct format.' % (base, mn))
                continue
            try:
                subnum = int(fn[2])
            except ValueError:
                print('%s/submissions/%s: filename does not follow correct format.' % (base, mn))
                continue

            # check for supplementary data file
            supfiles = get_files_in_dir(base + '/submissions', targname.split('.')[0] + '_supplementary')
            if len(supfiles) > 0:
                sup_url = 'submissions/%s' % supfiles[0]
                shutil.copyfile(base + '/submissions/' + supfiles[0], static + '/submissions/' + supfiles[0])
            else:
                sup_url = ""   # zip if more than one file

            sub_data = extract_data_from_sub(base + '/submissions/' + sub)
            submission_details[sub_data['Submission']['submission_id']] = {'url': 'submissions/%s' % targname, 'sup_url': sup_url, 'date': sub_data['Submission']['submission_date'], 'filename': targname, 'data': sub_data}
            shutil.copyfile(base + '/submissions/' + sub, static + '/submissions/' + targname)

def extract_data_from_sequences():
    global sequence_details

    makempty(static + '/sequences')

    # Filename format must be 'inferred_sequence_<number>.xlsx'
    seqs = get_files_in_dir(base + '/sequences', '.xlsx')
    for seq in seqs:
        fn = seq.replace(' ', '_')
        while('__') in fn:
            fn = fn.replace('__', '_')
        targname = fn.lower()
        fn = fn.replace('.xlsx', '')
        fn = fn.split('_')
        if len(fn) != 3:
            print('%s/sequences/%s: filename does not follow correct format.' % (base, mn))
            continue
        try:
            subnum = int(fn[2])
        except ValueError:
            print('%s/sequences/%s: filename does not follow correct format.' % (base, mn))
            continue

        seq_data = extract_data_from_seq(base + '/sequences/' + seq, minute_details)
        sequence_details[seq_data['Sequence']['description_id']] = {'url': 'sequences/%s' % targname, 'date': seq_data['Sequence']['release_date'], 'filename': targname, 'data': seq_data, 'affirmation': seq_data['Sequence']['affirmation_level']}
        shutil.copyfile(base + '/sequences/' + seq, static + '/sequences/' + targname)


def extract_data_from_sub(filename):
    wb = load_workbook(filename)
    data = {}
    data['Submission'] = extract_vert_table(wb, 'Submission', 'Field', 'Response', filename)
    data['Acknowledgements'] = extract_horiz_table(wb, 'Submission', 'name', filename)
    data['Repertoire'] = extract_vert_table(wb, 'Repertoire', 'Field', 'Response', filename)
    data['Inferences'] = extract_horiz_table(wb, 'Inferences', 'sequence_id', filename)
    data['Inference_list'] = list(data['Inferences'].keys())
    data['Genotype'] = extract_horiz_table(wb, 'Genotype', 'sequence_id', filename)
    data['Genotype_list'] = list(data['Genotype'].keys())
    data['Tool_Settings'] = extract_vert_table(wb, 'Tool Settings', 'Field', 'Response', filename)

    return data

# Extract tabulated data from an inferred sequence sheet
def extract_data_from_seq(filename, minute_details):
    wb = load_workbook(filename)
    data = {}
    data['Sequence'] = extract_vert_table(wb, 'Sequence', 'Field', 'Response', filename)
    data['Acknowledgements'] = extract_horiz_table(wb, 'Sequence', 'name', filename)
#    data['Delineation'] = extract_vert_tables(wb, 'Delineation', 'Field', 'Response', filename)
    data['Submissions'] = extract_vert_table(wb, 'Submissions', 'Submission ID', 'Sequence ID', filename)
    data['Meetings'] = extract_list(wb, 'Meetings', 'Meeting Number', filename)
    return data


# Relatively generic functions to extract data from Excel tables

# Extract contents of a table from the named tab
# Tables are assumed to start in column B
# Keys to the table run down the first column
# 'keyname' is the topmost key in the column
# desired responses are in the column whose topmost key is 'respname'
# result is a dictionary of key:response pairs
# filename is used for printing errors only
def extract_vert_table(wb, tabname, keyname, respname, filename):
    if tabname not in wb:
        print('No tab %s in workbook %s' % (tabname, filename))
        return {}

    ws = wb[tabname]
    res = OrderedDict()
    resp_col = None
    recording = False

    for cell in ws['B']:
        if recording:
            if cell.value == None:
                break
            else:
                res[cell.value] = ws.cell(row=cell.row, column=resp_col).value
        elif cell.value != None and keyname in cell.value:
            for rcell in ws[cell.row]:
                if rcell.value != None and respname in rcell.value:
                    resp_col = column_index_from_string(rcell.column)
                    break
            if resp_col != None:
                recording = True

    if len(res) < 1:
        print("Could not find a table with keyname %s and response name %s on tab %s of workbook %s" % (keyname, respname, tabname, filename))

    return res

# Extract multiple vertical tables from the named tab
# Tables are assumed to start in column B
# The table name should immediately precede the table
# Keys to the table run down the first column
# 'keyname' is the topmost key in the column
# desired responses are in the column whose topmost key is 'respname'
# result is a set of dictionaries of key:response pairs
# filename is used for printing errors only
def extract_vert_tables(wb, tabname, keyname, respname, filename):
    if tabname not in wb:
        print('No tab %s in workbook %s' % (tabname, filename))
        return {}

    ws = wb[tabname]
    res = OrderedDict()
    table = None
    resp_col = None
    recording = False
    header = None

    for cell in ws['B']:
        if recording:
            if cell.value == None:
                res[header] = table
                table = None
                header = None
                recording = False
            else:
                table[cell.value] = ws.cell(row=cell.row, column=resp_col).value
        elif cell.value != None:
            if (header != None) and (keyname in cell.value):
                for rcell in ws[cell.row]:
                    if rcell.value != None and respname in rcell.value:
                        resp_col = column_index_from_string(rcell.column)
                        break
                if resp_col != None:
                    table = OrderedDict()
                    recording = True
            else:
                header = cell.value

    if recording:
        res[header] = table

    if len(res) < 1:
        print("Could not find table(s) with keyname %s and response name %s on tab %s of workbook %s" % (keyname, respname, tabname, filename))

    return res

# Extract contents of a list (single column) from the named tab
# The list must be in column B
# 'keyname' is the topmost item in the column (header)
# result is a list of strings
# filename is used for printing errors only
def extract_list(wb, tabname, keyname, filename):
    if tabname not in wb:
        print('No tab %s in workbook %s' % (tabname, filename))
        return {}

    ws = wb[tabname]
    res = []
    recording = False

    for cell in ws['B']:
        if recording:
            if cell.value == None:
                break
            else:
                res.append(cell.value)
        elif cell.value != None and keyname in cell.value:
            recording = True

    if len(res) < 1:
        print("Could not find a list with keyname %s on tab %s of workbook %s" % (keyname, tabname, filename))

    return res


# Extract contents of table(s) from the named tab
# Tables are assumed to start in column B
# Keys to the table run along the first row
# 'keyname' is the leftmost topmost key in the row
# The table is preceded by the cells 'Subject id' and 'Genotype Id'
# These index, potentially, multiple tables
# If they are blank, serially ascending numbers will be used
# result is a set of dictionaries of key:response pairs
# filename is used for printing errors only
def extract_horiz_table(wb, tabname, keyname, filename):
    if tabname not in wb:
        print('No tab %s in workbook %s' % (tabname, filename))
        return {}

    ws = wb[tabname]
    res = OrderedDict()
    sub_id = None
    gen_id = None
    sub_ind = itertools.count(1)
    gen_ind = itertools.count(1)
    recording = False
    keys = []

    for cell in ws['B']:
        if recording:
            if cell is None or cell.value is None or len(cell.value) == 0 or len(cell.value.replace(" ", '')) == 0:
                if len(table) > 0:
                    if sub_id and gen_id:
                        res['Sub_%s_Gen_%s' % (str(sub_id), str(gen_id))] = table
                    else:
                        res = table
                table = None
                sub_id = None
                gen_id = None
                keys = []
                recording = False
            else:
                rec = OrderedDict()
                for rcell in ws[cell.row]:
                    ind = column_index_from_string(rcell.column) - column_index_from_string(cell.column)
                    if ind >= 0:
                        if ind < len(keys):
                            if keys[ind] != None:
                                rec[keys[ind]] = rcell.value
                if len(rec) > 0:
                    table.append(rec)
        elif cell.value != None and 'subject id' in cell.value.lower():
            val = get_value_to_right(ws, cell)
            sub_id = val if val != None else next(sub_ind)
        elif cell.value != None and 'genotype id' in cell.value.lower():
            val = get_value_to_right(ws, cell)
            gen_id = val if val != None else next(gen_ind)
        elif cell.value != None and keyname in cell.value:
            keys = []
            for rcell in ws[cell.row]:
                if column_index_from_string(rcell.column) >= column_index_from_string(cell.column):
                    keys.append(rcell.value)
            if len(keys) > 0:
                table = []
                recording = True
        else:
            sub_id = None
            gen_id = None

    if len(res) < 1:
        print("Could not find a table with keyname %s on tab %s of workbook %s" % (keyname, tabname, filename))

    return res

def get_value_to_right(ws, cell):
    return ws.cell(row=cell.row, column=column_index_from_string(cell.column) + 1).value

# File manipulation utilities

def get_immediate_subdirectories(a_dir):
    return [name for name in os.listdir(a_dir)
            if os.path.isdir(os.path.join(a_dir, name))]

def get_files_in_dir(a_dir, match):
    return [name for name in os.listdir(a_dir)
            if (not os.path.isdir(os.path.join(a_dir, name)) )
               and match in name]

# because rmtree doesn't work on Windows...
def deltree(path):
    for root, dirs, files in os.walk(path, topdown=False):
        for name in files:
            os.remove(os.path.join(root, name))
        for name in dirs:
            os.rmdir(os.path.join(root, name))

def makempty(dir):
    if os.path.isdir(dir):
        deltree(dir)
    if not os.path.isdir(dir):
        os.mkdir(dir)


if __name__ == "__main__":
    main(sys.argv)
